"""Excel I/O — reads regions, writes results.

Column layout (1-based):
  Col 1  : Locale
  Col 2  : URL
  Then for each check, 5 columns:
  Col 3  : Test Name
  Col 4  : Expected
  Col 5  : Actual
  Col 6  : Translated
  Col 7  : Result
  ... repeats for next check
"""

from typing import List, Optional

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

from .config_loader import GlobalConfig, SuiteConfig, CheckDef
from .check_runner import CheckResult

# ── Styles ────────────────────────────────────────────────────────────────────
_FILL_PASS    = PatternFill("solid", start_color="C6EFCE")
_FILL_FAIL    = PatternFill("solid", start_color="FFC7CE")
_FILL_SKIP    = PatternFill("solid", start_color="EFEFEF")
_FILL_HDR     = PatternFill("solid", start_color="1e1e2e")
_FONT_PASS    = Font(name="Arial", size=10, bold=True, color="276221")
_FONT_FAIL    = Font(name="Arial", size=10, bold=True, color="9C0006")
_FONT_SKIP    = Font(name="Arial", size=10, italic=True, color="888888")
_FONT_HDR     = Font(name="Arial", size=10, bold=True, color="FFFFFF")
_FONT_NORMAL  = Font(name="Arial", size=10)
_WRAP         = Alignment(wrap_text=True, vertical="top")

# Fixed columns
_COL_LOCALE = 1
_COL_URL    = 2
_COLS_PER_CHECK = 5   # test name | expected | actual | translated | result


def _check_start_col(check_index: int) -> int:
    return 3 + check_index * _COLS_PER_CHECK


def load_regions(global_cfg: GlobalConfig, suite: SuiteConfig) -> list[dict]:
    wb = load_workbook(suite.excel_file)
    ws = wb.active

    if not ws.cell(1, 1).value:
        _write_headers(ws, suite)
        wb.save(suite.excel_file)

    regions = []
    for row in ws.iter_rows(min_row=2):
        val = row[0].value
        if val:
            regions.append({"row": row[0].row, "region": str(val).strip()})

    wb.close()
    return regions


def write_results(
    global_cfg: GlobalConfig,
    suite: SuiteConfig,
    row: int,
    region: str,
    url: str,
    results: List[Optional[CheckResult]],
) -> None:
    wb = load_workbook(suite.excel_file)
    ws = wb.active

    # Locale + URL
    _cell(ws, row, _COL_LOCALE, region, _FONT_NORMAL)
    _cell(ws, row, _COL_URL,    url,    _FONT_NORMAL)

    for i, (check, result) in enumerate(zip(suite.checks, results)):
        col = _check_start_col(i)

        # Test name
        _cell(ws, row, col,     check.name,       _FONT_NORMAL)

        # Expected
        _cell(ws, row, col + 1, check.expected.strip(), _FONT_NORMAL)

        if result is None:
            # Skipped
            _cell(ws, row, col + 2, "N/A",      _FONT_SKIP)
            _cell(ws, row, col + 3, "N/A",      _FONT_SKIP)
            c = ws.cell(row=row, column=col + 4)
            c.value, c.fill, c.font = "SKIPPED", _FILL_SKIP, _FONT_SKIP
        else:
            # Actual
            _cell(ws, row, col + 2, result.actual_text, _FONT_NORMAL)
            # Translated
            _cell(ws, row, col + 3, result.translated or "—", _FONT_NORMAL)
            # Result
            c = ws.cell(row=row, column=col + 4)
            c.value = "PASS" if result.passed else "FAIL"
            c.fill  = _FILL_PASS if result.passed else _FILL_FAIL
            c.font  = _FONT_PASS if result.passed else _FONT_FAIL

    wb.save(suite.excel_file)
    wb.close()


# ── Internal ──────────────────────────────────────────────────────────────────

def _cell(ws, row, col, value, font):
    c = ws.cell(row=row, column=col)
    c.value     = value
    c.font      = font
    c.alignment = _WRAP


def _write_headers(ws, suite: SuiteConfig) -> None:
    _hdr(ws, 1, _COL_LOCALE, "Locale")
    _hdr(ws, 1, _COL_URL,    "URL")

    for i, check in enumerate(suite.checks):
        col = _check_start_col(i)
        _hdr(ws, 1, col,     f"{check.name} — Test Name")
        _hdr(ws, 1, col + 1, f"{check.name} — Expected")
        _hdr(ws, 1, col + 2, f"{check.name} — Actual")
        _hdr(ws, 1, col + 3, f"{check.name} — Translated")
        _hdr(ws, 1, col + 4, f"{check.name} — Result")


def _hdr(ws, row, col, value):
    c = ws.cell(row=row, column=col)
    c.value     = value
    c.font      = _FONT_HDR
    c.fill      = _FILL_HDR
    c.alignment = _WRAP